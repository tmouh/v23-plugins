from __future__ import annotations

import os
import tempfile
import zipfile
from pathlib import Path
from typing import TypedDict


class InventoryEntry(TypedDict):
    source_path: str
    preview: str


MAX_PREVIEW_CHARS = 4000
MAX_CSV_PREVIEW_ROWS = 20
TEXT_SUFFIXES = {".csv", ".tsv", ".txt", ".md", ".json"}
IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}
PDF_SUFFIXES = {".pdf"}
EXCEL_SUFFIXES = {".xlsx", ".xls"}
WORD_SUFFIXES = {".docx", ".doc"}
ZIP_SUFFIXES = {".zip"}
IGNORE_SUFFIXES = {".ds_store", ".ini", ".lnk"}


def _preview_text(path: Path) -> str:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return ""
    if path.suffix.lower() in {".csv", ".tsv"}:
        lines = text.splitlines()[:MAX_CSV_PREVIEW_ROWS]
        return "\n".join(lines)[:MAX_PREVIEW_CHARS]
    return text[:MAX_PREVIEW_CHARS]


def _preview_image(path: Path) -> str:
    from PIL import Image
    try:
        with Image.open(path) as im:
            w, h = im.size
            return f"image {path.suffix.lower().lstrip('.')} {w}x{h}"
    except Exception as e:
        return f"image (unreadable: {e})"


def _preview_pdf(path: Path) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            if not pdf.pages:
                return "PDF (empty)"
            text = pdf.pages[0].extract_text() or ""
        return text[:MAX_PREVIEW_CHARS] or "PDF (no extractable text on first page)"
    except Exception as e:
        return f"PDF (preview failed: {e})"


def _preview_excel(path: Path) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = []
        for sheet_name in wb.sheetnames[:1]:
            ws = wb[sheet_name]
            lines.append(f"Sheet: {sheet_name}")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= MAX_CSV_PREVIEW_ROWS:
                    break
                lines.append(",".join("" if v is None else str(v) for v in row))
        return "\n".join(lines)[:MAX_PREVIEW_CHARS]
    except Exception as e:
        return f"Excel (preview failed: {e})"


def _preview_word(path: Path) -> str:
    # Defer Word support — return a placeholder the LLM can still work with.
    return f"Word document {path.name}"


def _preview_binary(path: Path) -> str:
    size = path.stat().st_size
    return f"binary {path.suffix.lower().lstrip('.')} {size} bytes"


def _safe_extract_zip(zip_path: Path, out_dir: Path) -> None:
    """Extract zip_path into out_dir atomically, rejecting zip-slip members.

    Extraction goes to a sibling temp directory first, then renames into
    place so a crashed mid-extraction leaves no partial <name>__extracted folder.
    """
    out_dir = out_dir.resolve()
    parent = out_dir.parent
    parent.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory(dir=parent, prefix=f".{out_dir.name}.tmp-") as tmp:
        tmp_path = Path(tmp)
        with zipfile.ZipFile(zip_path, "r") as zf:
            for member in zf.infolist():
                # Reject absolute paths and parent-dir traversal.
                dest = (tmp_path / member.filename).resolve()
                # On Windows, use os.sep; also cover the case where dest == tmp_path
                # (e.g. an empty filename).
                try:
                    dest.relative_to(tmp_path)
                except ValueError:
                    raise ValueError(
                        f"Zip slip detected in {zip_path.name}: member {member.filename!r} "
                        f"escapes extraction directory"
                    )
            zf.extractall(tmp_path)

        # Atomically move the fully-extracted dir into its final name.
        # TemporaryDirectory will try to clean up after the with-block, but
        # the path will no longer exist — that's fine (its __exit__ handles missing dirs).
        os.replace(tmp_path, out_dir)


def _preview_for(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in TEXT_SUFFIXES:
        return _preview_text(path)
    if suffix in IMAGE_SUFFIXES:
        return _preview_image(path)
    if suffix in PDF_SUFFIXES:
        return _preview_pdf(path)
    if suffix in EXCEL_SUFFIXES:
        return _preview_excel(path)
    if suffix in WORD_SUFFIXES:
        return _preview_word(path)
    return _preview_binary(path)


def scan_folder(root: Path) -> list[InventoryEntry]:
    """Walk `root` recursively, extract a preview for each regular file.

    Auto-extracts .zip archives (and nested zips inside them) into sibling
    `<name>__extracted` folders so all contents show up in the inventory.
    """
    root = Path(root).resolve()
    if not root.exists():
        return []

    # Extract zips until none remain unprocessed (handles nested zips).
    while True:
        pending = [
            p for p in root.rglob("*")
            if p.is_file()
            and p.suffix.lower() in ZIP_SUFFIXES
            and not (p.parent / f"{p.stem}__extracted").exists()
        ]
        if not pending:
            break
        for p in pending:
            target = p.parent / f"{p.stem}__extracted"
            _safe_extract_zip(p, target)

    items: list[InventoryEntry] = []
    for p in root.rglob("*"):
        if not p.is_file():
            continue
        if p.suffix.lower() in IGNORE_SUFFIXES:
            continue
        if p.suffix.lower() in ZIP_SUFFIXES:
            continue
        items.append(InventoryEntry(
            source_path=str(p.resolve()),
            preview=_preview_for(p),
        ))
    items.sort(key=lambda it: it["source_path"])
    return items
