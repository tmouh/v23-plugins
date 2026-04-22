from __future__ import annotations

import csv
import hashlib
import json
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl

from deal_pack.models import RentRollSummary, T12Summary, Derived


SIDECAR_NAME = ".v23-deal-pack.sha"


def _dec(value) -> float | str:
    """Convert Decimal to a float for xlsx writing (openpyxl handles Decimal, but floats render more naturally)."""
    if isinstance(value, Decimal):
        return float(value)
    return value


def _write_csv_as_sheet(wb: openpyxl.Workbook, sheet_name: str, csv_path: Path) -> None:
    ws = wb.create_sheet(title=sheet_name[:31])  # xlsx max 31 chars
    with open(csv_path, newline="", encoding="utf-8") as f:
        for row in csv.reader(f):
            ws.append(row)


def _write_rent_roll_summary(wb: openpyxl.Workbook, s: RentRollSummary) -> None:
    ws = wb.create_sheet(title="rent_roll_summary")
    ws.append(["metric", "value"])
    ws.append(["total_sqft", s.total_sqft])
    ws.append(["occupied_sqft", s.occupied_sqft])
    ws.append(["vacancy_pct", _dec(s.vacancy_pct)])
    ws.append(["walt_years", _dec(s.walt_years)])
    ws.append(["avg_base_rent_psf", _dec(s.avg_base_rent_psf)])
    ws.append(["total_annual_base_rent", _dec(s.total_annual_base_rent)])


def _write_t12_summary(wb: openpyxl.Workbook, s: T12Summary) -> None:
    ws = wb.create_sheet(title="t12_summary")
    ws.append(["metric", "value"])
    ws.append(["gross_revenue", _dec(s.gross_revenue)])
    ws.append(["effective_gross_income", _dec(s.effective_gross_income)])
    ws.append(["opex_total", _dec(s.opex_total)])
    ws.append(["opex_per_sqft", _dec(s.opex_per_sqft)])
    ws.append(["noi", _dec(s.noi)])
    ws.append(["noi_per_sqft", _dec(s.noi_per_sqft)])


def _write_derived(wb: openpyxl.Workbook, d: Derived) -> None:
    ws = wb.create_sheet(title="derived")
    ws.append(["metric", "value"])
    ws.append(["noi_margin", _dec(d.noi_margin)])
    ws.append([
        "implied_cap_rate_at_ask",
        _dec(d.implied_cap_rate_at_ask) if d.implied_cap_rate_at_ask is not None else "",
    ])
    ws.append(["expense_ratio", _dec(d.expense_ratio)])


def write_financials_xlsx(
    *,
    path: Path,
    rent_roll_rows_csv: Optional[Path],
    rent_roll_summary: RentRollSummary,
    t12_rows_csv: Optional[Path],
    t12_summary: T12Summary,
    derived: Derived,
    raw_rent_roll_csv: Optional[Path] = None,
    raw_t12_csv: Optional[Path] = None,
) -> None:
    """Write multi-tab financials.xlsx.

    Tabs:
      - rent_roll        (if rent_roll_rows_csv provided)
      - rent_roll_summary
      - t12              (if t12_rows_csv provided)
      - t12_summary
      - derived
      - raw_rent_roll    (if raw_rent_roll_csv provided — low confidence)
      - raw_t12          (if raw_t12_csv provided — low confidence)
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    if rent_roll_rows_csv is not None:
        _write_csv_as_sheet(wb, "rent_roll", rent_roll_rows_csv)
    _write_rent_roll_summary(wb, rent_roll_summary)

    if t12_rows_csv is not None:
        _write_csv_as_sheet(wb, "t12", t12_rows_csv)
    _write_t12_summary(wb, t12_summary)

    _write_derived(wb, derived)

    if raw_rent_roll_csv is not None:
        _write_csv_as_sheet(wb, "raw_rent_roll", raw_rent_roll_csv)
    if raw_t12_csv is not None:
        _write_csv_as_sheet(wb, "raw_t12", raw_t12_csv)

    wb.save(path)


def write_manifest(path: Path, manifest: dict) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8",
    )


def _sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    h.update(path.read_bytes())
    return h.hexdigest()


def _read_sidecar(sidecar: Path) -> dict:
    """Load sidecar JSON, returning {} if missing, empty, or corrupted.

    Corrupted or partial writes (OneDrive sync, antivirus interruption) must not
    crash the skill — treat them as 'no recorded hashes' and let the next write
    replace the sidecar cleanly.
    """
    if not sidecar.exists():
        return {}
    try:
        raw = sidecar.read_text(encoding="utf-8")
        if not raw.strip():
            return {}
        return json.loads(raw)
    except (json.JSONDecodeError, OSError):
        return {}


def write_facts_sidecar(facts_path: Path) -> None:
    """Record the current hash of facts.md so future re-runs can tell if the user edited it."""
    facts_path = Path(facts_path)
    sidecar = facts_path.parent / SIDECAR_NAME
    if not facts_path.exists():
        return
    existing = _read_sidecar(sidecar)
    existing[facts_path.name] = _sha256_of(facts_path)
    sidecar.write_text(json.dumps(existing, indent=2), encoding="utf-8")


def is_facts_modified_by_user(facts_path: Path) -> bool:
    facts_path = Path(facts_path)
    sidecar = facts_path.parent / SIDECAR_NAME
    if not facts_path.exists():
        return False
    recorded = _read_sidecar(sidecar).get(facts_path.name)
    if recorded is None:
        return False
    return _sha256_of(facts_path) != recorded
