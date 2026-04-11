"""V23 Placement Engine — XLSX exporter for formatted placement lists.

Generates a formatted Excel file from ranked investor data (JSON array).
No database dependency — takes pre-built investor dicts and outputs xlsx.
"""

import argparse
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_COLUMNS = [
    ("Status", 12),
    ("Cov.", 6),
    ("Capital Group", 30),
    ("Contact", 25),
    ("Email", 30),
    ("Last", 12),
    ("OM", 12),
    ("Placement Comments", 40),
    ("Match Notes", 50),
]

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center")

_TIER_CONFIG = {
    1: ("Tier 1 \u2014 Strong Match", "E2EFDA"),
    2: ("Tier 2 \u2014 Possible Match", "FFF2CC"),
    3: ("Tier 3 \u2014 Long Shot", "FCE4D6"),
}

_NUM_COLS = len(_COLUMNS)  # 9


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def export_placement_list(ranked_investors, output_path, deal_name=None):
    """Generate a formatted placement list xlsx.

    Args:
        ranked_investors: list of dicts with keys: investor_name, coverage_owner,
            contact_name, email, match_notes, tier (1/2/3).
        output_path: path to write the .xlsx file.
        deal_name: optional deal name for the header row.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Placement List"

    # -- Column widths --
    for col_idx, (_, width) in enumerate(_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # -- Row 1: Deal header --
    if deal_name:
        ws.merge_cells("A1:I1")
        cell = ws["A1"]
        cell.value = f"Deal: {deal_name}"
        cell.font = Font(bold=True, size=14)

    # -- Row 3: Column headers --
    for col_idx, (name, _) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=3, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    # -- Data rows starting at row 4 --
    current_row = 4
    current_tier = None

    for inv in ranked_investors:
        tier = inv.get("tier", 1)

        # Insert tier separator when tier changes
        if tier != current_tier:
            current_tier = tier
            label, color_hex = _TIER_CONFIG.get(tier, (f"Tier {tier}", "DDDDDD"))
            tier_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            tier_font = Font(bold=True, size=11)

            merge_range = f"A{current_row}:I{current_row}"
            ws.merge_cells(merge_range)
            cell = ws.cell(row=current_row, column=1, value=label)
            cell.font = tier_font
            cell.fill = tier_fill
            # Apply fill to all cells in the merged range
            for c in range(2, _NUM_COLS + 1):
                ws.cell(row=current_row, column=c).fill = tier_fill
            current_row += 1

        # Write investor row
        tier_fill = PatternFill(
            start_color=_TIER_CONFIG.get(tier, ("", "DDDDDD"))[1],
            end_color=_TIER_CONFIG.get(tier, ("", "DDDDDD"))[1],
            fill_type="solid",
        )

        row_data = [
            None,                          # Status (blank)
            inv.get("coverage_owner", ""),  # Cov.
            inv.get("investor_name", ""),   # Capital Group
            inv.get("contact_name", ""),    # Contact
            inv.get("email", ""),           # Email
            None,                          # Last (blank)
            None,                          # OM (blank)
            None,                          # Placement Comments (blank)
            inv.get("match_notes", ""),     # Match Notes
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.fill = tier_fill

        current_row += 1

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="V23 Placement Engine — XLSX Exporter",
    )
    parser.add_argument(
        "--input", required=True,
        help="JSON file with ranked investors array",
    )
    parser.add_argument(
        "--output", required=True,
        help="Output .xlsx file path",
    )
    parser.add_argument(
        "--deal-name", default=None,
        help="Optional deal name for header row",
    )
    return parser


def main():
    parser = _build_parser()
    args = parser.parse_args()

    with open(args.input, "r") as f:
        ranked_investors = json.load(f)

    export_placement_list(ranked_investors, args.output, deal_name=args.deal_name)

    result = {
        "status": "ok",
        "output": args.output,
        "count": len(ranked_investors),
    }
    print(json.dumps(result))


if __name__ == "__main__":
    main()
