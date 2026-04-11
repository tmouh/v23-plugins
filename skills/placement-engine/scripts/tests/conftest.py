"""Shared pytest fixtures for V23 Placement Engine tests."""

import pytest
import tempfile
import os
import sys
import openpyxl

# Add scripts directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


@pytest.fixture
def tmp_dir():
    """Temporary directory, cleaned up after each test."""
    with tempfile.TemporaryDirectory() as d:
        yield d


@pytest.fixture
def tmp_db(tmp_dir):
    """Path for a temporary test database (file does not exist yet)."""
    return os.path.join(tmp_dir, "test.db")


@pytest.fixture
def seeded_db(tmp_db):
    """Database with 3 investors and 2 deals pre-loaded. No interactions."""
    from db import create_database, insert_deal, insert_investor

    create_database(tmp_db)
    insert_deal(
        tmp_db, "Test Deal Alpha", deal_date="2026-01-01",
        asset_class="multifamily", geography="Tampa, FL",
        strategy="value-add", capital_stack_position="LP equity",
    )
    insert_deal(
        tmp_db, "Test Deal Beta", deal_date="2025-06-15",
        asset_class="industrial", geography="Dallas, TX",
        strategy="ground-up", capital_stack_position="LP equity",
    )
    insert_investor(tmp_db, "Acme Capital", aliases=["Acme"], coverage_owner="HC")
    insert_investor(tmp_db, "Beta Partners", aliases=["Beta LP"], coverage_owner="MS")
    insert_investor(tmp_db, "Gamma Group", coverage_owner="SM")
    return tmp_db


@pytest.fixture
def format_a_xlsx(tmp_dir):
    """Sample Format A placement xlsx (Capital Group columns)."""
    path = os.path.join(tmp_dir, "format_a.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detail"

    # Deal header in row 1
    ws["A1"] = "Deal: Test Deal Alpha  01-Jan-26"

    # Column headers in row 3
    headers = [
        "Row #", "Status", "Cov.", "Capital Group", "Contact", "Email",
        "Contact - Notes", "New Contact", "New - Role", "New - Email",
        "Date - Last", "Placement Comments", "Old Comments",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Data rows starting at row 4
    data = [
        [1, "Pass", "HC", "Acme Capital", "John Smith", "john@acme.com",
         "", "", "", "", "2025-12-01", "Not interested in this market",
         "Previous pass on similar deal"],
        [2, "Reviewing", "MS", "Beta Partners", "Jane Doe", "jane@beta.com",
         "", "", "", "", "2026-01-15", "Reviewing terms", ""],
        [3, "Sent", "SM", "Gamma Group", "Bob Wilson", "bob@gamma.com",
         "", "Alice Brown", "VP", "alice@gamma.com", "", "", ""],
    ]
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(path)
    return path


@pytest.fixture
def format_b_xlsx(tmp_dir):
    """Sample Format B placement xlsx (Capital Provider columns)."""
    path = os.path.join(tmp_dir, "format_b.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detail"

    ws["A1"] = "Deal: Test Deal Beta  15-Feb-26"

    headers = [
        "Status", "Coverage", "Capital Provider", "Contact Person",
        "Contact Email", "Contact Numbers", "Date Sent",
        "Placement Comments", "Previous / Other Commentary",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    data = [
        ["1.Actively Reviewing", "HC", "Delta Fund", "Tom Lee",
         "tom@delta.com", "555-0001", "2026-01-20",
         "Very interested", "Met at conference"],
        ["5.Pass", "MS", "Epsilon LLC", "Sara Chen",
         "sara@epsilon.com", "555-0002", "2026-01-10",
         "Too small", ""],
        ["2.Reviewing", "SM", "Zeta Investments", "Mike Park",
         "mike@zeta.com", "555-0003", "2026-01-25",
         "Under review by IC", "Previous interest"],
    ]
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(path)
    return path
