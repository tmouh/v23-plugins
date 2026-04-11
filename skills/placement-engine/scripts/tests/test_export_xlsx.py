"""Tests for export_xlsx — V23 Placement Engine XLSX exporter.

TDD: these tests are written first, then export_xlsx.py is implemented.
"""

import json
import os
import subprocess
import sys

import openpyxl
import pytest

from export_xlsx import export_placement_list


# ── Fixtures ──────────────────────────────────────────────────────────


@pytest.fixture
def sample_investors():
    """Ranked investor list spanning all three tiers."""
    return [
        {
            "investor_name": "Acme Capital",
            "coverage_owner": "HC",
            "contact_name": "John Smith",
            "email": "john@acme.com",
            "match_notes": "Strong multifamily track record in FL",
            "tier": 1,
        },
        {
            "investor_name": "Beta Partners",
            "coverage_owner": "MS",
            "contact_name": "Jane Doe",
            "email": "jane@beta.com",
            "match_notes": "Interested in value-add, but prefers TX",
            "tier": 2,
        },
        {
            "investor_name": "Gamma Group",
            "coverage_owner": "SM",
            "contact_name": "Bob Wilson",
            "email": "bob@gamma.com",
            "match_notes": "Expanding into new markets",
            "tier": 2,
        },
        {
            "investor_name": "Delta Fund",
            "coverage_owner": "HC",
            "contact_name": "Tom Lee",
            "email": "tom@delta.com",
            "match_notes": "Long shot, different strategy focus",
            "tier": 3,
        },
    ]


@pytest.fixture
def single_tier_investors():
    """Investors all in the same tier (no tier separator needed between rows)."""
    return [
        {
            "investor_name": "Alpha LLC",
            "coverage_owner": "HC",
            "contact_name": "Alice A",
            "email": "alice@alpha.com",
            "match_notes": "Perfect match",
            "tier": 1,
        },
        {
            "investor_name": "Bravo Inc",
            "coverage_owner": "MS",
            "contact_name": "Bob B",
            "email": "bob@bravo.com",
            "match_notes": "Also strong",
            "tier": 1,
        },
    ]


# ── File creation ─────────────────────────────────────────────────────


class TestFileCreation:
    def test_creates_file_at_output_path(self, tmp_dir, sample_investors):
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        assert os.path.exists(path)

    def test_file_is_valid_xlsx(self, tmp_dir, sample_investors):
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        assert len(wb.sheetnames) >= 1
        wb.close()


# ── Column headers ────────────────────────────────────────────────────


class TestColumnHeaders:
    def test_correct_9_column_headers_in_row_3(self, tmp_dir, sample_investors):
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        expected = [
            "Status", "Cov.", "Capital Group", "Contact", "Email",
            "Last", "OM", "Placement Comments", "Match Notes",
        ]
        actual = [ws.cell(row=3, column=c).value for c in range(1, 10)]
        assert actual == expected
        wb.close()

    def test_header_row_formatting(self, tmp_dir, sample_investors):
        """Column headers should have blue background and white bold font."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        cell = ws.cell(row=3, column=1)
        # Check blue fill (openpyxl may store with 00 or FF alpha prefix)
        fill_rgb = cell.fill.start_color.rgb
        assert fill_rgb in ("FF4472C4", "004472C4"), f"Expected blue fill, got {fill_rgb}"
        # Check white bold font
        assert cell.font.bold is True
        font_rgb = cell.font.color.rgb
        assert font_rgb in ("FFFFFFFF", "00FFFFFF"), f"Expected white font, got {font_rgb}"
        wb.close()


# ── Deal header ───────────────────────────────────────────────────────


class TestDealHeader:
    def test_deal_name_in_header_row(self, tmp_dir, sample_investors):
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path, deal_name="Test Deal Alpha")
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        assert ws["A1"].value == "Deal: Test Deal Alpha"
        wb.close()

    def test_deal_header_bold_size_14(self, tmp_dir, sample_investors):
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path, deal_name="Test Deal Alpha")
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        cell = ws["A1"]
        assert cell.font.bold is True
        assert cell.font.size == 14
        wb.close()

    def test_deal_header_merged(self, tmp_dir, sample_investors):
        """A1:I1 should be merged for the deal header."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path, deal_name="Test Deal Alpha")
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        merged = [str(m) for m in ws.merged_cells.ranges]
        assert any("A1" in m and "I1" in m for m in merged)
        wb.close()

    def test_no_deal_header_when_none(self, tmp_dir, sample_investors):
        """When deal_name is None, A1 should be empty or None."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        assert ws["A1"].value is None or ws["A1"].value == ""
        wb.close()


# ── Investor data rows ────────────────────────────────────────────────


class TestInvestorData:
    def test_investor_name_in_capital_group_column(self, tmp_dir, sample_investors):
        """Capital Group (col C / col 3) should contain investor_name."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        # First data row: row 4 is tier separator for tier 1,
        # row 5 is first investor
        # Find first non-separator data row
        investor_names = []
        for row in ws.iter_rows(min_row=4, max_col=9, values_only=False):
            val = row[2].value  # Column C (0-indexed: col 3)
            if val and not str(val).startswith("Tier"):
                investor_names.append(val)

        assert "Acme Capital" in investor_names
        assert "Beta Partners" in investor_names
        assert "Gamma Group" in investor_names
        assert "Delta Fund" in investor_names
        wb.close()

    def test_match_notes_in_last_column(self, tmp_dir, sample_investors):
        """Match Notes (col I / col 9) should contain match_notes."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        match_notes = []
        for row in ws.iter_rows(min_row=4, max_col=9, values_only=False):
            val = row[8].value  # Column I (0-indexed: col 9)
            if val:
                match_notes.append(val)

        assert "Strong multifamily track record in FL" in match_notes
        assert "Long shot, different strategy focus" in match_notes
        wb.close()

    def test_coverage_owner_in_cov_column(self, tmp_dir, sample_investors):
        """Cov. (col B / col 2) should contain coverage_owner."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        cov_values = []
        for row in ws.iter_rows(min_row=4, max_col=9, values_only=False):
            val = row[1].value  # Column B
            if val and not str(val).startswith("Tier"):
                cov_values.append(val)

        assert "HC" in cov_values
        assert "MS" in cov_values
        assert "SM" in cov_values
        wb.close()

    def test_contact_and_email_columns(self, tmp_dir, sample_investors):
        """Contact (col D) and Email (col E) should be populated."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        contacts = []
        emails = []
        for row in ws.iter_rows(min_row=4, max_col=9, values_only=False):
            c = row[3].value  # Column D
            e = row[4].value  # Column E
            if c:
                contacts.append(c)
            if e:
                emails.append(e)

        assert "John Smith" in contacts
        assert "john@acme.com" in emails
        wb.close()

    def test_status_last_om_comments_blank(self, tmp_dir, sample_investors):
        """Status, Last, OM, Placement Comments should be blank for new outreach."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        for row in ws.iter_rows(min_row=4, max_col=9, values_only=False):
            # Skip tier separator rows
            if row[2].value and str(row[2].value).startswith("Tier"):
                continue
            # Only check investor rows (have a Capital Group value)
            if row[2].value:
                assert row[0].value is None or row[0].value == ""  # Status (A)
                assert row[5].value is None or row[5].value == ""  # Last (F)
                assert row[6].value is None or row[6].value == ""  # OM (G)
                assert row[7].value is None or row[7].value == ""  # Comments (H)
        wb.close()


# ── Tier separators ───────────────────────────────────────────────────


class TestTierSeparators:
    def test_tier_separator_rows_present(self, tmp_dir, sample_investors):
        """When tiers change, separator rows should be inserted."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        separator_texts = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            val = row[0]
            if val and isinstance(val, str) and val.startswith("Tier"):
                separator_texts.append(val)

        assert any("Tier 1" in t for t in separator_texts)
        assert any("Tier 2" in t for t in separator_texts)
        assert any("Tier 3" in t for t in separator_texts)
        wb.close()

    def test_tier_separator_labels(self, tmp_dir, sample_investors):
        """Tier separators should have exact label text."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        separator_texts = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            val = row[0]
            if val and isinstance(val, str) and val.startswith("Tier"):
                separator_texts.append(val)

        # Use \u2014 for em dash
        assert "Tier 1 \u2014 Strong Match" in separator_texts
        assert "Tier 2 \u2014 Possible Match" in separator_texts
        assert "Tier 3 \u2014 Long Shot" in separator_texts
        wb.close()

    def test_no_duplicate_tier_separator(self, tmp_dir, sample_investors):
        """Tier 2 separator should appear only once even with 2 tier-2 investors."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        tier2_count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            val = row[0]
            if val and isinstance(val, str) and "Tier 2" in val:
                tier2_count += 1

        assert tier2_count == 1
        wb.close()

    def test_single_tier_has_one_separator(self, tmp_dir, single_tier_investors):
        """All investors in same tier should produce exactly one separator."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(single_tier_investors, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        separator_count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            val = row[0]
            if val and isinstance(val, str) and val.startswith("Tier"):
                separator_count += 1

        assert separator_count == 1
        wb.close()

    def test_tier_separator_merged(self, tmp_dir, sample_investors):
        """Tier separator rows should have A:I merged."""
        path = os.path.join(tmp_dir, "output.xlsx")
        export_placement_list(sample_investors, path, deal_name="Merge Test")
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        all_merges = [str(m) for m in ws.merged_cells.ranges]
        # 1 deal header (A1:I1) + 3 tier separators = 4 merged ranges
        assert len(all_merges) >= 4, f"Expected >= 4 merges, got {len(all_merges)}: {all_merges}"
        # Verify at least one tier merge exists beyond row 1
        tier_merges = [m for m in all_merges if not m.startswith("A1")]
        assert len(tier_merges) >= 3
        wb.close()


# ── CLI ───────────────────────────────────────────────────────────────


class TestCLI:
    """Test the CLI via subprocess."""

    @property
    def _script(self):
        return os.path.join(
            os.path.dirname(__file__), "..", "export_xlsx.py"
        )

    def test_cli_creates_file_and_outputs_json(self, tmp_dir, sample_investors):
        input_path = os.path.join(tmp_dir, "investors.json")
        output_path = os.path.join(tmp_dir, "placement.xlsx")

        with open(input_path, "w") as f:
            json.dump(sample_investors, f)

        result = subprocess.run(
            [
                sys.executable, self._script,
                "--input", input_path,
                "--output", output_path,
                "--deal-name", "CLI Test Deal",
            ],
            capture_output=True, text=True,
        )

        assert result.returncode == 0
        data = json.loads(result.stdout)
        assert data["status"] == "ok"
        assert data["output"] == output_path
        assert data["count"] == 4
        assert os.path.exists(output_path)

    def test_cli_without_deal_name(self, tmp_dir, sample_investors):
        input_path = os.path.join(tmp_dir, "investors.json")
        output_path = os.path.join(tmp_dir, "no_deal.xlsx")

        with open(input_path, "w") as f:
            json.dump(sample_investors, f)

        result = subprocess.run(
            [
                sys.executable, self._script,
                "--input", input_path,
                "--output", output_path,
            ],
            capture_output=True, text=True,
        )

        assert result.returncode == 0
        data = json.loads(result.stdout)
        assert data["status"] == "ok"
        assert os.path.exists(output_path)
