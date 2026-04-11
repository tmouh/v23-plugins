"""Tests for parse_xlsx — V23 Placement Engine XLSX parser.

TDD: these tests are written first, then parse_xlsx.py is implemented.
"""

import json
import os
import subprocess
import sys

import openpyxl
import pytest

from parse_xlsx import (
    detect_format,
    extract_deal_header,
    find_header_row,
    parse_format_a,
    parse_format_b,
    parse_placement_file,
)


# ── detect_format ──────────────────────────────────────────────────────


class TestDetectFormat:
    def test_format_a(self, format_a_xlsx):
        assert detect_format(format_a_xlsx) == "A"

    def test_format_b(self, format_b_xlsx):
        assert detect_format(format_b_xlsx) == "B"

    def test_edge_unknown_columns(self, tmp_dir):
        """File with no known marker columns should return 'edge'."""
        path = os.path.join(tmp_dir, "unknown.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Random"
        ws["B1"] = "Spreadsheet"
        wb.save(path)
        assert detect_format(path) == "edge"

    def test_format_a_header_not_in_first_row(self, tmp_dir):
        """Capital Group header in row 5 (within first 10) should be detected."""
        path = os.path.join(tmp_dir, "a_deep.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=5, column=3, value="Capital Group")
        wb.save(path)
        assert detect_format(path) == "A"

    def test_format_b_on_second_sheet(self, tmp_dir):
        """Capital Provider on a non-first sheet should still detect as B."""
        path = os.path.join(tmp_dir, "b_sheet2.xlsx")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Summary"
        ws2 = wb.create_sheet("Detail")
        ws2.cell(row=3, column=1, value="Capital Provider")
        wb.save(path)
        assert detect_format(path) == "B"


# ── find_header_row ────────────────────────────────────────────────────


class TestFindHeaderRow:
    def test_finds_header_row_format_a(self, format_a_xlsx):
        wb = openpyxl.load_workbook(format_a_xlsx, read_only=True, data_only=True)
        ws = wb["Detail"]
        target_cols = ["Status", "Capital Group", "Contact", "Email"]
        row_num, col_map = find_header_row(ws, target_cols)
        assert row_num == 3
        assert "Status" in col_map
        assert "Capital Group" in col_map
        wb.close()

    def test_finds_header_row_format_b(self, format_b_xlsx):
        wb = openpyxl.load_workbook(format_b_xlsx, read_only=True, data_only=True)
        ws = wb["Detail"]
        target_cols = ["Status", "Capital Provider", "Contact Person"]
        row_num, col_map = find_header_row(ws, target_cols)
        assert row_num == 3
        assert "Capital Provider" in col_map
        wb.close()

    def test_returns_none_when_not_found(self, tmp_dir):
        path = os.path.join(tmp_dir, "empty.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Nothing"
        wb.save(path)
        wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws2 = wb2[wb2.sheetnames[0]]
        row_num, col_map = find_header_row(ws2, ["Capital Group", "Status"])
        assert row_num is None
        assert col_map is None
        wb2.close()

    def test_col_map_has_correct_indices(self, format_a_xlsx):
        """Column indices should allow correct cell lookup."""
        wb = openpyxl.load_workbook(format_a_xlsx, read_only=True, data_only=True)
        ws = wb["Detail"]
        target_cols = ["Status", "Capital Group"]
        row_num, col_map = find_header_row(ws, target_cols)
        # "Status" is column 2 (1-indexed), "Capital Group" is column 4
        assert col_map["Status"] == 2
        assert col_map["Capital Group"] == 4
        wb.close()


# ── extract_deal_header ────────────────────────────────────────────────


class TestExtractDealHeader:
    def test_format_a_deal_header(self, format_a_xlsx):
        wb = openpyxl.load_workbook(format_a_xlsx, read_only=True, data_only=True)
        ws = wb["Detail"]
        header = extract_deal_header(ws)
        assert header["deal_name"] == "Test Deal Alpha"
        assert header["deal_date"] == "01-Jan-26"
        wb.close()

    def test_format_b_deal_header(self, format_b_xlsx):
        wb = openpyxl.load_workbook(format_b_xlsx, read_only=True, data_only=True)
        ws = wb["Detail"]
        header = extract_deal_header(ws)
        assert header["deal_name"] == "Test Deal Beta"
        assert header["deal_date"] == "15-Feb-26"
        wb.close()

    def test_no_deal_header(self, tmp_dir):
        """Sheet with no Deal: header returns Nones."""
        path = os.path.join(tmp_dir, "no_header.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Not a deal header"
        wb.save(path)
        wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws2 = wb2[wb2.sheetnames[0]]
        header = extract_deal_header(ws2)
        assert header["deal_name"] is None
        assert header["deal_date"] is None
        wb2.close()


# ── parse_format_a ─────────────────────────────────────────────────────


class TestParseFormatA:
    def test_returns_correct_row_count(self, format_a_xlsx):
        result = parse_format_a(format_a_xlsx)
        assert len(result["rows"]) == 3

    def test_deal_header_present(self, format_a_xlsx):
        result = parse_format_a(format_a_xlsx)
        assert result["deal_header"]["deal_name"] == "Test Deal Alpha"

    def test_investor_names(self, format_a_xlsx):
        result = parse_format_a(format_a_xlsx)
        names = [r["investor_name"] for r in result["rows"]]
        assert names == ["Acme Capital", "Beta Partners", "Gamma Group"]

    def test_status_mapping(self, format_a_xlsx):
        result = parse_format_a(format_a_xlsx)
        statuses = [r["status"] for r in result["rows"]]
        assert statuses == ["Pass", "Reviewing", "Sent"]

    def test_coverage_code_mapping(self, format_a_xlsx):
        result = parse_format_a(format_a_xlsx)
        codes = [r["coverage_code"] for r in result["rows"]]
        assert codes == ["HC", "MS", "SM"]

    def test_contact_fields(self, format_a_xlsx):
        result = parse_format_a(format_a_xlsx)
        row0 = result["rows"][0]
        assert row0["contact_name"] == "John Smith"
        assert row0["email"] == "john@acme.com"

    def test_new_contact_fields(self, format_a_xlsx):
        result = parse_format_a(format_a_xlsx)
        row2 = result["rows"][2]  # Gamma Group has new contact
        assert row2["new_contact"] == "Alice Brown"
        assert row2["new_contact_role"] == "VP"
        assert row2["new_contact_email"] == "alice@gamma.com"

    def test_comments_mapping(self, format_a_xlsx):
        result = parse_format_a(format_a_xlsx)
        assert result["rows"][0]["raw_comments"] == "Not interested in this market"
        assert result["rows"][0]["old_comments"] == "Previous pass on similar deal"

    def test_date_last_contact(self, format_a_xlsx):
        result = parse_format_a(format_a_xlsx)
        assert result["rows"][0]["date_last_contact"] == "2025-12-01"

    def test_skips_empty_investor_rows(self, tmp_dir):
        """Rows with no investor_name should be skipped."""
        path = os.path.join(tmp_dir, "sparse_a.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detail"
        ws["A1"] = "Deal: Sparse Deal  01-Jan-26"
        headers = ["Row #", "Status", "Cov.", "Capital Group", "Contact"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        # Row with investor
        ws.cell(row=4, column=2, value="Pass")
        ws.cell(row=4, column=4, value="Real Investor")
        # Row without investor (empty)
        ws.cell(row=5, column=2, value="Reviewing")
        # Row without investor (None-like)
        ws.cell(row=6, column=4, value=None)
        wb.save(path)
        result = parse_format_a(path)
        assert len(result["rows"]) == 1
        assert result["rows"][0]["investor_name"] == "Real Investor"

    def test_none_string_normalized(self, tmp_dir):
        """String 'None' from openpyxl should become actual None."""
        path = os.path.join(tmp_dir, "none_str_a.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detail"
        ws["A1"] = "Deal: None Test  01-Jan-26"
        headers = ["Row #", "Status", "Cov.", "Capital Group", "Contact", "Email"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        ws.cell(row=4, column=4, value="Test Investor")
        ws.cell(row=4, column=5, value="None")
        ws.cell(row=4, column=6, value="None")
        wb.save(path)
        result = parse_format_a(path)
        assert result["rows"][0]["contact_name"] is None
        assert result["rows"][0]["email"] is None

    def test_prefers_detail_sheet(self, tmp_dir):
        """Should use 'Detail' sheet when available."""
        path = os.path.join(tmp_dir, "multi_sheet_a.xlsx")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Summary"
        ws1["A1"] = "Ignore this"

        ws2 = wb.create_sheet("Detail")
        ws2["A1"] = "Deal: Multi Sheet Deal  01-Jan-26"
        headers = ["Row #", "Status", "Cov.", "Capital Group"]
        for col, h in enumerate(headers, 1):
            ws2.cell(row=3, column=col, value=h)
        ws2.cell(row=4, column=4, value="Sheet2 Investor")
        ws2.cell(row=4, column=2, value="Pass")
        wb.save(path)

        result = parse_format_a(path)
        assert result["rows"][0]["investor_name"] == "Sheet2 Investor"


# ── parse_format_b ─────────────────────────────────────────────────────


class TestParseFormatB:
    def test_returns_correct_row_count(self, format_b_xlsx):
        result = parse_format_b(format_b_xlsx)
        assert len(result["rows"]) == 3

    def test_deal_header_present(self, format_b_xlsx):
        result = parse_format_b(format_b_xlsx)
        assert result["deal_header"]["deal_name"] == "Test Deal Beta"

    def test_investor_names(self, format_b_xlsx):
        result = parse_format_b(format_b_xlsx)
        names = [r["investor_name"] for r in result["rows"]]
        assert names == ["Delta Fund", "Epsilon LLC", "Zeta Investments"]

    def test_status_strips_numeric_prefix(self, format_b_xlsx):
        """'1.Actively Reviewing' -> 'Actively Reviewing', '5.Pass' -> 'Pass'."""
        result = parse_format_b(format_b_xlsx)
        statuses = [r["status"] for r in result["rows"]]
        assert statuses == ["Actively Reviewing", "Pass", "Reviewing"]

    def test_coverage_code(self, format_b_xlsx):
        result = parse_format_b(format_b_xlsx)
        codes = [r["coverage_code"] for r in result["rows"]]
        assert codes == ["HC", "MS", "SM"]

    def test_contact_fields(self, format_b_xlsx):
        result = parse_format_b(format_b_xlsx)
        row0 = result["rows"][0]
        assert row0["contact_name"] == "Tom Lee"
        assert row0["email"] == "tom@delta.com"
        assert row0["phone"] == "555-0001"

    def test_comments_mapping(self, format_b_xlsx):
        result = parse_format_b(format_b_xlsx)
        assert result["rows"][0]["raw_comments"] == "Very interested"
        assert result["rows"][0]["old_comments"] == "Met at conference"

    def test_date_last_contact(self, format_b_xlsx):
        result = parse_format_b(format_b_xlsx)
        assert result["rows"][0]["date_last_contact"] == "2026-01-20"

    def test_skips_empty_investor_rows(self, tmp_dir):
        """Rows with no investor_name should be skipped."""
        path = os.path.join(tmp_dir, "sparse_b.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Deal: Sparse B  01-Jan-26"
        headers = ["Status", "Coverage", "Capital Provider"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        ws.cell(row=4, column=1, value="1.Pass")
        ws.cell(row=4, column=3, value="Real Fund")
        ws.cell(row=5, column=1, value="2.Reviewing")
        # row 5 has no investor_name
        wb.save(path)
        result = parse_format_b(path)
        assert len(result["rows"]) == 1

    def test_status_prefix_variations(self, tmp_dir):
        """Various numeric prefix patterns should all be stripped."""
        path = os.path.join(tmp_dir, "prefix_b.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Deal: Prefix Test  01-Jan-26"
        headers = ["Status", "Coverage", "Capital Provider"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        statuses = ["1.Active", "5.Pass", "2.Reviewing", "10.Other"]
        for i, s in enumerate(statuses):
            ws.cell(row=4 + i, column=1, value=s)
            ws.cell(row=4 + i, column=3, value=f"Investor {i}")
        wb.save(path)
        result = parse_format_b(path)
        parsed_statuses = [r["status"] for r in result["rows"]]
        assert parsed_statuses == ["Active", "Pass", "Reviewing", "Other"]


# ── parse_placement_file (auto-detect) ─────────────────────────────────


class TestParsePlacementFile:
    def test_auto_detects_format_a(self, format_a_xlsx):
        result = parse_placement_file(format_a_xlsx)
        assert result["format"] == "A"
        assert len(result["rows"]) == 3
        assert result["source_file"] == format_a_xlsx

    def test_auto_detects_format_b(self, format_b_xlsx):
        result = parse_placement_file(format_b_xlsx)
        assert result["format"] == "B"
        assert len(result["rows"]) == 3
        assert result["source_file"] == format_b_xlsx

    def test_edge_format_returns_empty_rows(self, tmp_dir):
        """Unknown format should return empty rows with a note."""
        path = os.path.join(tmp_dir, "edge.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Random data"
        wb.save(path)
        result = parse_placement_file(path)
        assert result["format"] == "edge"
        assert result["rows"] == []
        assert "note" in result

    def test_result_has_deal_header(self, format_a_xlsx):
        result = parse_placement_file(format_a_xlsx)
        assert "deal_header" in result
        assert result["deal_header"]["deal_name"] == "Test Deal Alpha"


# ── CLI ────────────────────────────────────────────────────────────────


class TestCLI:
    """Test the CLI subcommands via subprocess."""

    @property
    def _script(self):
        return os.path.join(
            os.path.dirname(__file__), "..", "parse_xlsx.py"
        )

    def test_cli_detect_format_a(self, format_a_xlsx):
        result = subprocess.run(
            [sys.executable, self._script, "detect", format_a_xlsx],
            capture_output=True, text=True,
        )
        assert result.returncode == 0
        data = json.loads(result.stdout)
        assert data["format"] == "A"

    def test_cli_detect_format_b(self, format_b_xlsx):
        result = subprocess.run(
            [sys.executable, self._script, "detect", format_b_xlsx],
            capture_output=True, text=True,
        )
        assert result.returncode == 0
        data = json.loads(result.stdout)
        assert data["format"] == "B"

    def test_cli_parse_format_a(self, format_a_xlsx):
        result = subprocess.run(
            [sys.executable, self._script, "parse", format_a_xlsx],
            capture_output=True, text=True,
        )
        assert result.returncode == 0
        data = json.loads(result.stdout)
        assert data["format"] == "A"
        assert len(data["rows"]) == 3

    def test_cli_parse_format_b(self, format_b_xlsx):
        result = subprocess.run(
            [sys.executable, self._script, "parse", format_b_xlsx],
            capture_output=True, text=True,
        )
        assert result.returncode == 0
        data = json.loads(result.stdout)
        assert data["format"] == "B"
        assert len(data["rows"]) == 3
