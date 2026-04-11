"""V23 Placement Engine — XLSX parser for commercial real estate placement lists.

Supports two formats:
  - Format A: "Capital Group" column style
  - Format B: "Capital Provider" column style

Provides auto-detection, header/row extraction, and a CLI.
"""

import argparse
import json
import os
import re
import sys

import openpyxl


# ---------------------------------------------------------------------------
# 1. detect_format
# ---------------------------------------------------------------------------

def detect_format(file_path: str) -> str:
    """Detect placement file format by scanning column headers.

    Checks the first 10 rows across all sheets for marker column names.
    Returns "A" if "Capital Group" found, "B" if "Capital Provider" found,
    or "edge" if neither is found.
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        for ws in wb:
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str):
                        stripped = val.strip()
                        if stripped == "Capital Group":
                            return "A"
                        if stripped == "Capital Provider":
                            return "B"
        return "edge"
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 2. find_header_row
# ---------------------------------------------------------------------------

def find_header_row(ws, target_columns: list):
    """Search the first 15 rows for a row with >= 2 matching column names.

    Args:
        ws: An openpyxl worksheet.
        target_columns: List of column name strings to look for.

    Returns:
        (row_number, {col_name: col_index}) if found, or (None, None).
        col_index is 1-based.
    """
    # Map lowercase -> canonical form so col_map keys match _FORMAT_*_MAP keys
    canonical = {c.strip().lower(): c.strip() for c in target_columns}

    for row in ws.iter_rows(min_row=1, max_row=15, values_only=False):
        col_map = {}
        for cell in row:
            if not hasattr(cell, 'column'):
                continue
            val = cell.value
            if isinstance(val, str) and val.strip().lower() in canonical:
                col_map[canonical[val.strip().lower()]] = cell.column
        if len(col_map) >= 2:
            return (row[0].row, col_map)

    return (None, None)


# ---------------------------------------------------------------------------
# 3. extract_deal_header
# ---------------------------------------------------------------------------

def extract_deal_header(ws) -> dict:
    """Extract deal name and date from the first 5 rows of a worksheet.

    Looks for a cell matching "Deal: <name>  <date>" (double-space separator).
    Returns {"deal_name": ..., "deal_date": ...} with None for missing parts.
    """
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for val in row:
            if isinstance(val, str) and val.strip().startswith("Deal:"):
                text = val.strip()[len("Deal:"):].strip()
                # Split on double-space
                parts = text.split("  ", 1)
                deal_name = parts[0].strip() if parts else None
                deal_date = parts[1].strip() if len(parts) > 1 else None
                return {"deal_name": deal_name, "deal_date": deal_date}
    return {"deal_name": None, "deal_date": None}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize_value(val):
    """Convert openpyxl 'None' strings to actual None, strip strings."""
    if val is None:
        return None
    if isinstance(val, str):
        stripped = val.strip()
        if stripped == "None" or stripped == "":
            return None
        return stripped
    return val


def _get_sheet(wb, prefer="Detail"):
    """Return the preferred sheet if it exists, otherwise the first sheet."""
    if prefer in wb.sheetnames:
        return wb[prefer]
    return wb[wb.sheetnames[0]]


def _strip_numeric_prefix(status_str):
    """Strip leading numeric prefix like '1.', '5.', '10.' from status strings."""
    if status_str is None:
        return None
    if isinstance(status_str, str):
        match = re.match(r"^\d+\.\s*", status_str)
        if match:
            return status_str[match.end():]
    return status_str


# ---------------------------------------------------------------------------
# Format A column mappings
# ---------------------------------------------------------------------------

_FORMAT_A_COLUMNS = [
    "Status", "Cov.", "Capital Group", "Contact", "Email",
    "New Contact", "New - Role", "New - Email",
    "Date - Last", "Last",  # "Date - Last" or "Last" for date_last_contact
    "Placement Comments", "Old Comments", "OM",
]

_FORMAT_A_MAP = {
    "Status": "status",
    "Cov.": "coverage_code",
    "Capital Group": "investor_name",
    "Contact": "contact_name",
    "Email": "email",
    "New Contact": "new_contact",
    "New - Role": "new_contact_role",
    "New - Email": "new_contact_email",
    "Date - Last": "date_last_contact",
    "Last": "date_last_contact",
    "Placement Comments": "raw_comments",
    "Old Comments": "old_comments",
    "OM": "date_om_sent",
}


# ---------------------------------------------------------------------------
# 4. parse_format_a
# ---------------------------------------------------------------------------

def parse_format_a(file_path: str) -> dict:
    """Parse a Format A placement file.

    Returns {"deal_header": {...}, "rows": [{...}, ...]}
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = _get_sheet(wb, prefer="Detail")
        deal_header = extract_deal_header(ws)

        row_num, col_map = find_header_row(ws, _FORMAT_A_COLUMNS)
        if row_num is None:
            return {"deal_header": deal_header, "rows": []}

        # Build reverse map: col_index -> output field name
        idx_to_field = {}
        for col_name, col_idx in col_map.items():
            field = _FORMAT_A_MAP.get(col_name)
            if field and field not in idx_to_field.values():
                idx_to_field[col_idx] = field
            elif field:
                # Duplicate field (e.g., "Last" and "Date - Last" both -> date_last_contact)
                # Only add if not already mapped
                if field not in [v for v in idx_to_field.values()]:
                    idx_to_field[col_idx] = field

        rows = []
        for row in ws.iter_rows(min_row=row_num + 1, values_only=False):
            record = {}
            for cell in row:
                if not hasattr(cell, 'column'):
                    continue
                field = idx_to_field.get(cell.column)
                if field:
                    record[field] = _normalize_value(cell.value)

            # Skip rows with no investor_name
            if not record.get("investor_name"):
                continue

            rows.append(record)

        return {"deal_header": deal_header, "rows": rows}
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Format B column mappings
# ---------------------------------------------------------------------------

_FORMAT_B_COLUMNS = [
    "Status", "Coverage", "Capital Provider", "Contact Person",
    "Contact Email", "Contact Numbers", "Date Sent",
    "Placement Comments", "Previous / Other Commentary",
]

_FORMAT_B_MAP = {
    "Status": "status",
    "Coverage": "coverage_code",
    "Capital Provider": "investor_name",
    "Contact Person": "contact_name",
    "Contact Email": "email",
    "Contact Numbers": "phone",
    "Date Sent": "date_last_contact",
    "Placement Comments": "raw_comments",
    "Previous / Other Commentary": "old_comments",
}


# ---------------------------------------------------------------------------
# 5. parse_format_b
# ---------------------------------------------------------------------------

def parse_format_b(file_path: str) -> dict:
    """Parse a Format B placement file.

    Status values have numeric prefixes stripped (e.g., '1.Actively Reviewing' -> 'Actively Reviewing').
    Returns {"deal_header": {...}, "rows": [{...}, ...]}
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = _get_sheet(wb, prefer="Detail")
        deal_header = extract_deal_header(ws)

        row_num, col_map = find_header_row(ws, _FORMAT_B_COLUMNS)
        if row_num is None:
            return {"deal_header": deal_header, "rows": []}

        # Build reverse map: col_index -> output field name
        idx_to_field = {}
        for col_name, col_idx in col_map.items():
            field = _FORMAT_B_MAP.get(col_name)
            if field:
                idx_to_field[col_idx] = field

        rows = []
        for row in ws.iter_rows(min_row=row_num + 1, values_only=False):
            record = {}
            for cell in row:
                if not hasattr(cell, 'column'):
                    continue
                field = idx_to_field.get(cell.column)
                if field:
                    record[field] = _normalize_value(cell.value)

            # Skip rows with no investor_name
            if not record.get("investor_name"):
                continue

            # Strip numeric prefix from status
            record["status"] = _strip_numeric_prefix(record.get("status"))

            rows.append(record)

        return {"deal_header": deal_header, "rows": rows}
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 6. parse_placement_file
# ---------------------------------------------------------------------------

def parse_placement_file(file_path: str) -> dict:
    """Auto-detect format and parse a placement file.

    Returns {format, deal_header, rows, source_file} or edge case with note.
    """
    fmt = detect_format(file_path)

    if fmt == "A":
        result = parse_format_a(file_path)
        return {
            "format": "A",
            "deal_header": result["deal_header"],
            "rows": result["rows"],
            "source_file": file_path,
        }
    elif fmt == "B":
        result = parse_format_b(file_path)
        return {
            "format": "B",
            "deal_header": result["deal_header"],
            "rows": result["rows"],
            "source_file": file_path,
        }
    else:
        return {
            "format": "edge",
            "deal_header": {"deal_name": None, "deal_date": None},
            "rows": [],
            "source_file": file_path,
            "note": "Unrecognized format — no 'Capital Group' or 'Capital Provider' column found.",
        }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="V23 Placement Engine XLSX Parser",
    )
    sub = parser.add_subparsers(dest="command")

    p_detect = sub.add_parser("detect", help="Detect file format (A, B, or edge)")
    p_detect.add_argument("file_path", help="Path to the .xlsx file")

    p_parse = sub.add_parser("parse", help="Parse a placement file to JSON")
    p_parse.add_argument("file_path", help="Path to the .xlsx file")

    return parser


def main():
    parser = _build_parser()
    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        sys.exit(1)

    if args.command == "detect":
        fmt = detect_format(args.file_path)
        print(json.dumps({"format": fmt}))

    elif args.command == "parse":
        result = parse_placement_file(args.file_path)
        print(json.dumps(result, default=str))


if __name__ == "__main__":
    main()
