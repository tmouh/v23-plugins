#!/usr/bin/env python3
"""
V23 Comp Exporter v2 — Takes merged comp JSON and creates a formatted Excel file.
Auto-detects asset type and uses the appropriate column layout:
  - Standard CRE (office, retail, industrial, multifamily)
  - Hospitality/Hotel (with RevPAR, ADR, occupancy, rooms/keys, price/key)
  - Lease comps

Usage:
  python3 export_comps.py --input /tmp/comp-search/merged.json \
    --output "research/Texas Hotel Comps.xlsx" --type both
"""

import argparse
import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# --- Column Definitions Per Asset Type ---

STANDARD_SALE_COLUMNS = [
    ("address", "Address", 35), ("submarket", "Submarket", 20),
    ("sf", "SF", 12), ("units", "Units", 10), ("price", "Price", 18),
    ("price_psf", "$/PSF", 12), ("cap_rate", "Cap Rate", 10), ("date", "Date", 14),
    ("buyer", "Buyer", 25), ("seller", "Seller", 25), ("property_class", "Class", 8),
    ("year_built", "Built", 8), ("year_reno", "Reno", 8),
    ("notes", "Notes", 30), ("_source_file", "Source File", 35),
]

HOSPITALITY_SALE_COLUMNS = [
    ("property_name", "Property Name", 35), ("address", "Address", 30),
    ("city", "City", 15), ("state", "State", 6), ("submarket", "Submarket", 22),
    ("asset_type", "Type", 25), ("brand", "Brand/Flag", 30),
    ("rooms_keys", "Rooms/Keys", 12), ("price", "Sale Price", 18),
    ("price_per_key", "Price/Key", 14), ("price_psf", "$/PSF", 12),
    ("cap_rate", "Cap Rate", 10), ("noi", "NOI", 15),
    ("total_revenue", "Total Revenue", 15),
    ("revpar", "RevPAR", 10), ("adr", "ADR", 10), ("occupancy", "Occupancy", 10),
    ("date", "Date", 14), ("buyer", "Buyer", 25), ("seller", "Seller", 25),
    ("year_built", "Built", 8),
    ("notes", "Notes", 40), ("_source_file", "Source File", 35),
]

LEASE_COLUMNS = [
    ("address", "Address", 35), ("submarket", "Submarket", 20),
    ("tenant", "Tenant", 25), ("sf", "SF", 12), ("rent_psf", "Rent $/SF", 12),
    ("rent_monthly", "Rent/Mo", 12), ("rent_annual", "Rent/Year", 14),
    ("date", "Date", 14), ("unit_type", "Type", 10), ("unit", "Unit", 10),
    ("term", "Term", 10), ("expiration", "Expiration", 14),
    ("notes", "Notes", 30), ("source", "Source", 20), ("_source_file", "Source File", 35),
]

MARKET_REPORT_COLUMNS = [
    ("property_name", "Report/Property", 40), ("city", "City", 15),
    ("state", "State", 6), ("asset_type", "Asset Type", 20),
    ("brand", "Brand/Source", 30), ("date", "Date", 14),
    ("notes", "Notes", 50), ("_source_file", "Source File", 40),
]

# --- Styling ---

HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
ALT_ROW_FILL = PatternFill(start_color="F5F8FA", end_color="F5F8FA", fill_type="solid")
DATA_FONT = Font(name="Arial", size=10)
SUMMARY_FONT = Font(name="Arial", size=10, bold=True)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1B3A5C")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def detect_asset_type(comps):
    """Detect if this is a hospitality dataset based on field presence."""
    hospitality_fields = {'rooms_keys', 'revpar', 'adr', 'price_per_key', 'brand'}
    hospitality_keywords = {'hotel', 'hospitality', 'resort', 'motel', 'inn',
                            'marriott', 'hilton', 'hyatt', 'select service'}
    hosp_score = 0
    for comp in comps:
        for field in hospitality_fields:
            if comp.get(field) is not None:
                hosp_score += 2
        for field in ['property_name', 'address', 'asset_type', 'notes', '_source_file']:
            val = str(comp.get(field, '')).lower()
            if any(kw in val for kw in hospitality_keywords):
                hosp_score += 3
    return "hospitality" if hosp_score > len(comps) else "standard"


def format_cell(ws, row, col, value, is_currency=False, is_pct=False, is_number=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if value is None:
        cell.alignment = CENTER
        return cell
    if is_currency and isinstance(value, (int, float)):
        cell.number_format = '$#,##0'
        cell.alignment = RIGHT
    elif is_pct and isinstance(value, (int, float)):
        cell.value = value / 100 if value >= 1 else value
        cell.number_format = '0.0%'
        cell.alignment = RIGHT
    elif is_number and isinstance(value, (int, float)):
        cell.number_format = '#,##0'
        cell.alignment = RIGHT
    elif isinstance(value, str) and len(value) > 40:
        cell.alignment = LEFT
    else:
        cell.alignment = LEFT
    return cell


def create_summary_sheet(wb, comps, comp_type, metadata=None):
    ws = wb.create_sheet("Summary", 0)
    ws.merge_cells('A1:D1')
    ws['A1'].value = f"V23 Comp Search Results — {comp_type.title()} Comps"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:D2')
    ws['A2'].value = f"Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ws['A2'].font = Font(name="Arial", size=10, color="888888", italic=True)

    row = 4
    stats = [("Total Comps", len(comps))]

    if metadata:
        stats.append(("Zones Searched", metadata.get("zones_searched", "N/A")))
        stats.append(("Files Scanned", metadata.get("total_files_searched", "N/A")))
        if metadata.get("files_cloud_only"):
            stats.append(("Cloud-Only Files (not read)", len(metadata["files_cloud_only"])))
        if metadata.get("files_encrypted"):
            stats.append(("Encrypted Files (not read)", len(metadata["files_encrypted"])))

    stats.append(("Source Files", len(set(c.get('_source_file', '') for c in comps))))

    sale_comps = [c for c in comps if c.get('_comp_type') == 'sale']
    lease_comps = [c for c in comps if c.get('_comp_type') == 'lease']
    market_reports = [c for c in comps if c.get('source_type') == 'market_report']

    if sale_comps and lease_comps:
        stats += [("Sale Comps", len(sale_comps)), ("Lease Comps", len(lease_comps))]
    if market_reports:
        stats.append(("Market Reports", len(market_reports)))

    # Sale stats
    if sale_comps:
        prices = [c['price'] for c in sale_comps if isinstance(c.get('price'), (int, float)) and c['price'] > 0]
        psfs = [c['price_psf'] for c in sale_comps if isinstance(c.get('price_psf'), (int, float)) and c['price_psf'] > 0]
        caps = [c['cap_rate'] for c in sale_comps if isinstance(c.get('cap_rate'), (int, float)) and 0 < c['cap_rate'] < 1]
        ppks = [c['price_per_key'] for c in sale_comps if isinstance(c.get('price_per_key'), (int, float)) and c['price_per_key'] > 0]
        adrs = [c['adr'] for c in sale_comps if isinstance(c.get('adr'), (int, float)) and c['adr'] > 0]
        occs = [c['occupancy'] for c in sale_comps if isinstance(c.get('occupancy'), (int, float)) and c['occupancy'] > 0]

        if prices:
            stats += [("Avg Sale Price", f"${sum(prices)/len(prices):,.0f}"),
                      ("Median Sale Price", f"${sorted(prices)[len(prices)//2]:,.0f}")]
        if psfs:
            stats += [("Avg $/PSF", f"${sum(psfs)/len(psfs):,.0f}")]
        if ppks:
            stats += [("Avg Price/Key", f"${sum(ppks)/len(ppks):,.0f}")]
        if caps:
            stats += [("Avg Cap Rate", f"{sum(caps)/len(caps):.2%}")]
        if adrs:
            stats += [("ADR Range", f"${min(adrs):,.2f} - ${max(adrs):,.2f}")]
        if occs:
            occ_vals = [o * 100 if o < 1 else o for o in occs]
            stats += [("Occupancy Range", f"{min(occ_vals):.1f}% - {max(occ_vals):.1f}%")]

    # Lease stats
    if lease_comps:
        rents = [c['rent_psf'] for c in lease_comps if isinstance(c.get('rent_psf'), (int, float)) and c['rent_psf'] > 0]
        if rents:
            stats += [("Avg Rent $/SF", f"${sum(rents)/len(rents):,.2f}")]

    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = SUMMARY_FONT
        ws.cell(row=row, column=2, value=value).font = DATA_FONT
        row += 1

    # Zone breakdown
    if metadata and metadata.get("zone_summary"):
        row += 1
        ws.cell(row=row, column=1, value="Results by Zone").font = SUMMARY_FONT
        row += 1
        for zone_info in metadata["zone_summary"]:
            ws.cell(row=row, column=1, value=zone_info["zone_name"]).font = DATA_FONT
            ws.cell(row=row, column=2, value=f"{zone_info['comps_found']} comps").font = DATA_FONT
            row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 30
    return ws


def write_data_sheet(wb, comps, sheet_name, columns):
    ws = wb.create_sheet(sheet_name)
    currency_cols = {'price', 'price_psf', 'rent_psf', 'rent_monthly', 'rent_annual',
                     'noi', 'total_revenue', 'price_per_key', 'revpar', 'adr'}
    pct_cols = {'cap_rate', 'occupancy'}
    number_cols = {'sf', 'units', 'year_built', 'year_reno', 'rooms_keys'}

    for col_idx, (key, label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, comp in enumerate(comps, 2):
        is_alt = row_idx % 2 == 0
        for col_idx, (key, label, width) in enumerate(columns, 1):
            cell = format_cell(ws, row_idx, col_idx, comp.get(key),
                               is_currency=(key in currency_cols),
                               is_pct=(key in pct_cols),
                               is_number=(key in number_cols))
            if is_alt:
                cell.fill = ALT_ROW_FILL

    ws.freeze_panes = "A2"
    if comps:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(comps) + 1}"
    return ws


def export_comps(data, output_path, comp_type="both"):
    comps = data.get('comps', [])
    metadata = {k: v for k, v in data.items() if k != 'comps'}

    if not comps:
        print("No comps to export.", file=sys.stderr)
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sale_comps = [c for c in comps if c.get('_comp_type') == 'sale']
    lease_comps = [c for c in comps if c.get('_comp_type') == 'lease']
    market_reports = [c for c in comps if c.get('source_type') == 'market_report']

    # Detect if hospitality
    asset_mode = detect_asset_type(sale_comps or comps)
    sale_columns = HOSPITALITY_SALE_COLUMNS if asset_mode == "hospitality" else STANDARD_SALE_COLUMNS

    create_summary_sheet(wb, comps, comp_type, metadata)

    if sale_comps:
        label = "Hotel Comps" if asset_mode == "hospitality" else "Sale Comps"
        write_data_sheet(wb, sale_comps, label, sale_columns)
    elif not lease_comps:
        # If no type classification, dump everything with detected columns
        label = "Hotel Comps" if asset_mode == "hospitality" else "Sale Comps"
        write_data_sheet(wb, comps, label, sale_columns)

    if lease_comps:
        write_data_sheet(wb, lease_comps, "Lease Comps", LEASE_COLUMNS)

    if market_reports:
        write_data_sheet(wb, market_reports, "Market Reports", MARKET_REPORT_COLUMNS)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    total = len(sale_comps) + len(lease_comps) + len(market_reports)
    print(f"Exported {total} comps ({asset_mode} mode) to: {output_path}", file=sys.stderr)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="V23 Comp Exporter v2")
    parser.add_argument('--input', required=True)
    parser.add_argument('--output', required=True)
    parser.add_argument('--type', choices=['sale', 'lease', 'both'], default='both')
    args = parser.parse_args()
    with open(args.input) as f:
        data = json.load(f)
    export_comps(data, args.output, args.type)


if __name__ == "__main__":
    main()
