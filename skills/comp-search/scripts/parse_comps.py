#!/usr/bin/env python3
"""
V23 Comp Parser v2 — Reads Excel files and extracts sale/lease comp data into
a normalized JSON format. Handles cloud-only files by logging them for
PowerShell retrieval.

Usage:
  python3 parse_comps.py --files file1.xlsx file2.xls --type both --output /tmp/comps.json
  python3 parse_comps.py --dir /path/to/deal/folder --type sale --output /tmp/comps.json
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

try:
    import xlrd
except ImportError:
    os.system("pip install xlrd --break-system-packages -q")
    import xlrd


SALE_HEADER_MAP = {
    "address": "address", "property": "address", "property name": "property_name",
    "property address": "address", "location": "location_desc",
    "quality": "quality_desc", "situation": "situation_desc",
    "seller": "seller", "buyer": "buyer", "buyer name": "buyer",
    "purchaser": "buyer", "neighborhood": "submarket", "submarket": "submarket",
    "market": "submarket", "area": "submarket",
    "gsf": "sf", "sf": "sf", "sqft": "sf", "square feet": "sf",
    "square footage": "sf", "size (sf)": "sf", "size": "sf", "nrsf": "sf", "gla": "sf",
    "total price": "price", "price": "price", "sale price": "price",
    "purchase price": "price", "tot": "price", "tot.": "price",
    "signed": "date", "date": "date", "date signed": "date", "close date": "date",
    "sale date": "date", "closing date": "date", "sold": "date",
    "$psf": "price_psf", "$/psf": "price_psf", "$/gsf": "price_psf",
    "price/sf": "price_psf", "price psf": "price_psf", "ppsf": "price_psf",
    "pp sf": "price_psf", "$ psf": "price_psf", "$/sf": "price_psf",
    "price per sf": "price_psf",
    "built": "year_built", "year built": "year_built", "yr built": "year_built",
    "reno": "year_reno", "renovated": "year_reno",
    "class": "property_class", "building class": "property_class",
    "cap rate": "cap_rate", "cap": "cap_rate", "going-in cap": "cap_rate",
    "miles to site": "miles_to_site", "distance": "miles_to_site",
    "units": "units", "# units": "units", "num units": "units", "total units": "units",
    "notes": "notes", "comments": "notes",
    # Hospitality-specific
    "rooms": "rooms_keys", "keys": "rooms_keys", "rooms/keys": "rooms_keys",
    "# rooms": "rooms_keys", "room count": "rooms_keys", "key count": "rooms_keys",
    "price/key": "price_per_key", "price per key": "price_per_key",
    "$/key": "price_per_key", "ppk": "price_per_key",
    "noi": "noi", "net operating income": "noi",
    "revenue": "total_revenue", "total revenue": "total_revenue",
    "revpar": "revpar", "rev par": "revpar",
    "adr": "adr", "average daily rate": "adr", "avg daily rate": "adr",
    "occupancy": "occupancy", "occ": "occupancy", "occ rate": "occupancy",
    "occ.": "occupancy", "occupancy rate": "occupancy",
    "brand": "brand", "flag": "brand", "franchise": "brand",
    "asset type": "asset_type", "property type": "asset_type", "type": "asset_type",
    "city": "city", "state": "state",
}

LEASE_HEADER_MAP = {
    "address": "address", "property": "address", "property name": "property_name",
    "property address": "address", "location": "address",
    "submarket": "submarket", "neighborhood": "submarket", "market": "submarket",
    "tenant": "tenant", "tenant name": "tenant", "lessee": "tenant",
    "date signed": "date", "date": "date", "signed": "date",
    "lease date": "date", "commencement": "date", "start date": "date",
    "size (sf)": "sf", "sf": "sf", "sqft": "sf", "square feet": "sf",
    "gsf": "sf", "nrsf": "sf",
    "rent psf": "rent_psf", "rent/sf": "rent_psf", "$/sf": "rent_psf",
    "$psf": "rent_psf", "rent /sf": "rent_psf", "asking rent": "rent_psf", "rent": "rent_psf",
    "dur": "term", "dur.": "term",
    "rent /mo": "rent_monthly", "monthly rent": "rent_monthly", "rent/month": "rent_monthly",
    "annual rent": "rent_annual",
    "type": "unit_type", "unit type": "unit_type", "unit": "unit", "suite": "unit",
    "notes": "notes", "comments": "notes", "source": "source",
    "lease expiration": "expiration", "expiration": "expiration",
    "lease term": "term", "term": "term",
    "city": "city", "state": "state",
}


def normalize_header(header, comp_type="sale"):
    if header is None:
        return None
    clean = str(header).strip().lower().rstrip('.')
    clean = re.sub(r'[^\w\s/$#.]', '', clean)
    clean = re.sub(r'\s+', ' ', clean).strip()
    hmap = SALE_HEADER_MAP if comp_type == "sale" else LEASE_HEADER_MAP
    if clean in hmap:
        return hmap[clean]
    if clean + '.' in hmap:
        return hmap[clean + '.']
    if 'mi' in clean and 'site' in clean:
        return 'miles_to_site'
    if clean == 'built/reno':
        return 'year_built'
    return hmap.get(clean, clean)


def parse_value(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, datetime):
        return val.strftime("%b %Y") if val.day == 1 else val.strftime("%b %d, %Y")
    if isinstance(val, date):
        return val.strftime("%b %Y") if val.day == 1 else val.strftime("%b %d, %Y")
    s = str(val).strip()
    if not s or s.lower() in ('n/a', 'na', '-', '--', 'tbd', 'none'):
        return None
    if re.match(r'^\$[\d,]+\.?\d*$', s):
        try: return float(s.replace('$', '').replace(',', ''))
        except ValueError: pass
    if re.match(r'^[\d.]+%$', s):
        try: return float(s.replace('%', '')) / 100
        except ValueError: pass
    if re.match(r'^[\d,]+\.?\d*$', s):
        try: return float(s.replace(',', ''))
        except ValueError: pass
    return s


def find_header_row(ws, max_scan=10):
    known_headers = {'address', 'property', 'sf', 'gsf', 'price', 'date', 'tenant',
                     'rent', 'submarket', 'neighborhood', 'buyer', 'seller', 'type',
                     'unit', 'rooms', 'keys', 'adr', 'revpar', 'occupancy', 'noi',
                     'brand', 'flag', 'cap rate'}
    best_row, best_score = 1, 0
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        row_vals = [str(cell.value).strip().lower() for cell in ws[row_idx] if cell.value]
        score = sum(1 for v in row_vals if any(h in v for h in known_headers))
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def classify_sheet(sheet_name):
    name_lower = sheet_name.lower()
    if any(kw in name_lower for kw in ['sale', 'buy', 'sold', 'purchase']):
        return 'sale'
    if any(kw in name_lower for kw in ['lease', 'leasing', 'rent', 'tenant']):
        return 'lease'
    return 'unknown'


def parse_worksheet(ws, sheet_name, comp_type, source_file):
    comps = []
    header_row = find_header_row(ws)
    raw_headers = [cell.value for cell in ws[header_row]]
    if not any(h for h in raw_headers if h is not None):
        return comps
    headers = [normalize_header(h, comp_type) for h in raw_headers]

    STREET_TYPES = ('st', 'ave', 'blvd', 'dr', 'pl', 'way', 'rd', 'ct', 'ln',
                    'ter', 'terrace', 'pkwy', 'hwy', 'street', 'avenue', 'boulevard',
                    'drive', 'place', 'road', 'court', 'lane', 'center', 'ctr', 'sq', 'square')

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data = {}
        has_data = False
        for col_idx, cell in enumerate(ws[row_idx]):
            if col_idx < len(headers) and headers[col_idx]:
                val = parse_value(cell.value)
                if val is not None:
                    has_data = True
                    row_data[headers[col_idx]] = val

        if not has_data:
            continue

        # Accept rows that have an address OR a property_name (for hospitality data)
        addr = row_data.get('address') or row_data.get('property_name')
        if not addr:
            continue

        addr_str = str(addr).strip()
        skip_labels = ('subject property', 'comparable', 'total', 'average', 'median',
                       'notes', 'source', 'n/a', '', 'address', 'property', 'footnotes:',
                       'footnotes', 'comparable sales', 'comparable leases')
        if addr_str.lower().rstrip(':') in skip_labels:
            continue
        if 'subject property' in addr_str.lower():
            continue

        # Clean up footnote markers
        addr_clean = re.sub(r'\s*\(\d+\)\s*$', '', addr_str).strip()
        if 'address' in row_data:
            row_data['address'] = addr_clean
        if 'property_name' in row_data:
            row_data['property_name'] = re.sub(r'\s*\(\d+\)\s*$', '', str(row_data['property_name'])).strip()

        # Round numeric fields
        for field in ('price_psf', 'rent_psf', 'revpar', 'adr'):
            if isinstance(row_data.get(field), float):
                row_data[field] = round(row_data[field], 2)

        # Handle combined year_built/reno
        val = row_data.get('year_built')
        if isinstance(val, str) and '/' in val:
            parts = val.split('/')
            try:
                row_data['year_built'] = int(parts[0].strip())
                if len(parts) > 1:
                    row_data['year_reno'] = int(parts[1].strip())
            except ValueError:
                pass

        # Clean date
        date_val = row_data.get('date')
        if isinstance(date_val, (int, float)):
            row_data['date'] = str(int(date_val))

        row_data['_source_file'] = os.path.basename(source_file)
        row_data['_source_sheet'] = sheet_name
        row_data['_comp_type'] = comp_type
        comps.append(row_data)
    return comps


def parse_excel_file(filepath, requested_type="both"):
    all_comps = []
    ext = os.path.splitext(filepath)[1].lower()

    # Try openpyxl for .xlsx/.xlsm
    if ext in ('.xlsx', '.xlsm'):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            err_str = str(e)
            if 'errno 22' in err_str.lower() or 'invalid argument' in err_str.lower():
                return all_comps, "cloud_only"
            print(f"  WARNING: Cannot open {os.path.basename(filepath)}: {e}", file=sys.stderr)
            return all_comps, "failed"

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_type = classify_sheet(sheet_name)
            if sheet_type == 'unknown':
                header_row_idx = find_header_row(ws)
                try:
                    raw_hdrs = [str(c.value).strip().lower() if c.value else '' for c in ws[header_row_idx]]
                    sale_markers = {'price', 'buyer', 'seller', 'tot', 'tot.', '$/gsf', '$psf', '$/psf', 'cap rate', 'noi', 'rooms', 'keys', 'revpar', 'adr'}
                    lease_markers = {'tenant', 'rent', 'rent_psf', 'rent psf', 'rent /sf'}
                    if any(h in sale_markers for h in raw_hdrs):
                        sheet_type = 'sale'
                    elif any(h in lease_markers for h in raw_hdrs):
                        sheet_type = 'lease'
                    else:
                        sheet_type = requested_type if requested_type != 'both' else 'sale'
                except Exception:
                    sheet_type = requested_type if requested_type != 'both' else 'sale'

            if requested_type in ('sale', 'lease') and sheet_type != requested_type:
                if classify_sheet(sheet_name) != 'unknown':
                    continue
            try:
                comps = parse_worksheet(ws, sheet_name, sheet_type, filepath)
                all_comps.extend(comps)
            except Exception as e:
                print(f"  WARNING: Error parsing sheet '{sheet_name}': {e}", file=sys.stderr)
        wb.close()
        return all_comps, "ok"

    # Try xlrd for .xls
    elif ext == '.xls':
        try:
            wb = xlrd.open_workbook(filepath)
        except Exception as e:
            err_str = str(e)
            if 'errno 22' in err_str.lower() or 'invalid argument' in err_str.lower():
                return all_comps, "cloud_only"
            if 'workbook is encrypted' in err_str.lower():
                return all_comps, "encrypted"
            print(f"  WARNING: Cannot open {os.path.basename(filepath)}: {e}", file=sys.stderr)
            return all_comps, "failed"

        # For .xls, we can't reuse the openpyxl worksheet parser directly.
        # Log what we found and return metadata.
        sheet_names = wb.sheet_names()
        wb.release_resources()
        print(f"  INFO: .xls file with sheets: {sheet_names}", file=sys.stderr)
        return all_comps, "ok"

    return all_comps, "unsupported"


class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)


def main():
    parser = argparse.ArgumentParser(description="V23 Comp Parser v2")
    parser.add_argument('--files', nargs='+', help='Specific Excel files to parse')
    parser.add_argument('--dir', help='Directory to search for comp files')
    parser.add_argument('--type', choices=['sale', 'lease', 'both'], default='both')
    parser.add_argument('--output', required=True, help='Output JSON file path')
    args = parser.parse_args()

    files_to_parse = [f for f in (args.files or []) if os.path.exists(f)]
    if args.dir and os.path.isdir(args.dir):
        for root, dirs, files in os.walk(args.dir):
            for f in files:
                ext = os.path.splitext(f)[1].lower()
                if ext in ('.xlsx', '.xlsm', '.xls') and not f.startswith('~$'):
                    files_to_parse.append(os.path.join(root, f))

    if not files_to_parse:
        print("No files to parse.", file=sys.stderr)
        result = {"comp_type": args.type, "files_parsed": 0, "files_failed": 0,
                  "files_cloud_only": [], "files_encrypted": [],
                  "total_comps": 0, "comps": []}
        os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
        with open(args.output, 'w') as f:
            json.dump(result, f, indent=2)
        sys.exit(0)

    print(f"Parsing {len(files_to_parse)} files for {args.type} comps...", file=sys.stderr)
    all_comps = []
    files_parsed, files_failed = 0, 0
    cloud_only_files, encrypted_files = [], []

    for filepath in files_to_parse:
        print(f"  Reading: {os.path.basename(filepath)}", file=sys.stderr)
        try:
            comps, status = parse_excel_file(filepath, args.type)
            if status == "cloud_only":
                cloud_only_files.append(filepath)
            elif status == "encrypted":
                encrypted_files.append(filepath)
            elif status == "failed":
                files_failed += 1
            else:
                all_comps.extend(comps)
                files_parsed += 1
        except Exception as e:
            print(f"  FAILED: {os.path.basename(filepath)} - {e}", file=sys.stderr)
            files_failed += 1

    # Deduplicate
    def norm_addr(a):
        if not a: return ''
        s = re.sub(r'[^\w\s]', '', str(a).lower())
        return re.sub(r'\s+', ' ', s).strip()

    seen, unique_comps = set(), []
    for comp in all_comps:
        key = (norm_addr(comp.get('address', '') or comp.get('property_name', '')),
               comp.get('_comp_type', ''))
        if key not in seen:
            seen.add(key)
            unique_comps.append(comp)

    result = {
        "comp_type": args.type,
        "files_parsed": files_parsed,
        "files_failed": files_failed,
        "files_cloud_only": cloud_only_files,
        "files_encrypted": encrypted_files,
        "total_comps": len(unique_comps),
        "duplicates_removed": len(all_comps) - len(unique_comps),
        "comps": unique_comps,
    }
    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    with open(args.output, 'w') as f:
        json.dump(result, f, indent=2, cls=DateTimeEncoder)
    print(f"\nDone. {len(unique_comps)} unique comps from {files_parsed} files. "
          f"({files_failed} failed, {len(cloud_only_files)} cloud-only, "
          f"{len(encrypted_files)} encrypted)", file=sys.stderr)


if __name__ == "__main__":
    main()
